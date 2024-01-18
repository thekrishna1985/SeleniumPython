# Pycharm move multiple lines left = Highlight lines + Shift + Tab
# Pycharm move multiple lines right = Highlight lines + Tab
# Pycharm Auto Indent = CTRL + ALT + I
# Pycharm Zoom In & Out (Default 12pts) = ALt+Shift+.(Increase), Alt+Shift+,(Decrease)
# ---------------------------------------------------------------------------------------------------------
# Import Webdriver Base Files + Additional Support Files
import logging
import time
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver import ActionChains
from selenium.webdriver.common import alert
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# Chrome Browser Requirements
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.service import Service as Chrome
from webdriver_manager.chrome import ChromeDriverManager

# Edge Browser Requirements
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.service import Service as Edge
from webdriver_manager.microsoft import EdgeChromiumDriverManager

# -----------------------------------------------------------------------------------------------------------
# Test With Chrome Browser
service = Chrome()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

# Test With Microsoft Edge Browser
# service = Edge()
# options = webdriver.EdgeOptions()
# driver = webdriver.Edge(service=service, options=options)

# -----------------------------------------------------------------------------------------------------------
# Open the Website Page & Basic Website Methods/Actions

# Sample Test Sites
# driver.get("https://demo.guru99.com/test/newtours/")
# driver.get("https://demo.automationtesting.in/Index.html")
# driver.get("https://rahulshettyacademy.com/seleniumPractise/#/")
# driver.get("https://rahulshettyacademy.com/dropdownsPractise/")
# driver.get("https://rahulshettyacademy.com/AutomationPractice/")
# driver.get("https://rahulshettyacademy.com/angularpractice/")
# driver.get("https://the-internet.herokuapp.com/")
# driver.get("https://omayo.blogspot.com/")

# Driver Implicitly Wait / Explicitly Wait
# driver.implicitly_wait(10)

# Window Size / Navigational Commands in Browser (Back/Forward Button in Browser)
# driver.maximize_window()
# driver.back()
# driver.forward()
# driver.refresh()

# Print Website Information
# print(driver.title)
# print(driver.current_url)

# -----------------------------------------------------------------------------------------------
# Using Keyboard buttons (Keys. (Will give the full list of keyboard actions))
# driver.get("https://rahulshettyacademy.com/angularpractice/")
# driver.find_element(By.XPATH, "//input[@name='bday']").send_keys(Keys.ARROW_DOWN, Keys.ENTER)
# time.sleep(2)
# ----------------------------------------------------------------------------------------------------------
# Finding, Selecting, Clicking on Elements
# Xpath & CSS Selector element can be selected on any element. (Name,Id,Class however depends on developers coding)
# Select only Single element by X-Path = Right click Element -> Inspect -> View in selectors hub --> Copy Rel X Path
# Or (Right-Click on element->Inspect->RightClick->Copy-CopyXpath (Change double quotes to single quotes)

# Clicking/Selecting Method 1 - (Normal Click Method)
# driver.find_element(By.XPATH, "XXXX").click()

# Clicking/Selecting Method 2 (Hidden behind other elements - Selenium element click intercepted exception) = Use Webdriverwait + Javascript

# wait = WebDriverWait(self.driver, 30)
# element = wait.until(EC.element_to_be_clickable((By.XPATH, "XXXX")))
# self.driver.execute_script("arguments[0].click()", element)

# Clicking/Selecting Method 3 - Use ActionChains (Hidden behind other elements - Selenium element click intercepted exception)

# path = driver.find_element(By.XPATH, "")
# action = ActionChains(driver)
# action.click(path).perform()

# -----------------------------------------------------------------------
# Selecting Single Element from Dynamic Lists = Select the blackberry from the Shopping list & Add to cart & Checkout

# driver.get("https://rahulshettyacademy.com/angularpractice/shop")
# Select the common class/container which holds all the items, in this case = //div[@class='card h-100']
# Put that in selector hub, it should highlight all elements
# Use and View the Selector Hub (Not HTML page) to travel down by using the "//" go the exact element you want and select the class/tag
# Due to chaining methods you don't need to use the full command, only input data after the "//"

# Loop through list items and select blackberry & add to cart
# product = driver.find_elements(By.XPATH, "//div[@class='card h-100']//h4//a")
#
# for i in product:
#     if i.text == "Blackberry":
#         i.find_element(By.XPATH, "/html/body/app-root/app-shop/div/div/div[2]/app-card-list/app-card[4]/div/div[2]/button").click()

# Click on checkout Button
# driver.find_element(By.XPATH, "//*[@id='navbarResponsive']/ul/li/a").click()
# time.sleep(5)

# -----------------------------------------------------------------------
# Selecting Multiple Elements From Dynamic List - Creating lists for items selected & pushing into list/array using Append
# Video on creating X Path - Lesson 48,52, - Functional automation example on Greenkart Application (Udemy Course)

# driver.get("https://rahulshettyacademy.com/seleniumPractise/#/")

# Search For Item
# driver.find_element(By.XPATH, "//*[@id='root']/div/header/div/div[2]/form/input").send_keys("ber")
# time.sleep(3)

# Create Xpath for all product elements = Look for main class holding all the products
# Main Class + 1 Level Down = //div[@class='products']/div
# results = driver.find_elements(By.XPATH, "//div[@class='products']/div")

# Create a list for the Items Searched, in this case "ber" & push items into the list using append (BONUS)
# expectedList = ["Cucumber - 1 Kg", "Raspberry - 1/4 Kg", "Strawberry - 1/4 Kg"]
# actualList = []

# Looping through list of items & Clicking add to Cart & push Items into Array list
# Add to cart X Path = travel down 1 div press enter. From all the elements there is only 1 element holding the button
# for result in results:
#     # Selecting the Item Names & Push into List
#     actualList.append(result.find_element(By.XPATH, "h4").text)
#
#     # Selecting the Add to Cart button
#     result.find_element(By.XPATH, "div/button").click()

# time.sleep(3)

# Verify if Name list Matches expected list
# print(expectedList)
# print(actualList)
# assert expectedList == actualList

# Selecting Check out Buttons
# driver.find_element(By.XPATH, "//*[@id='root']/div/header/div/div[3]/a[4]/img").click()
# driver.find_element(By.XPATH, "//*[@id='root']/div/header/div/div[3]/div[2]/div[2]/button").click()
# time.sleep(3)

# ---------------------------------------------------------------------------------------------------------------
# Selecting Dropdowns (Fixed/Dynamic)

# Fixed List Dropdown
# driver.get("https://rahulshettyacademy.com/angularpractice/")
# dropdown = Select(driver.find_element(By.XPATH, "//*[@id='exampleFormControlSelect1']"))
# dropdown.select_by_visible_text("Female")

# Hidden Drop-Down = Website Form = https://abx.com/become-a-member/ = https://stackoverflow.com/questions/76025965/selenium-dropdown-element-not-visible
# The SELECT you are targeting is hidden. Some sites create this non-SELECT dropdown that looks like a dropdown and then stores the value in a hidden SELECT.
# Because of this, you can't interact directly with the SELECT.
# For this site, you'll have to click the dropdown and then click the desired option from the dropdown. The code below will accomplish that.

# THE FIX = DROPDOWN 1 = COUNTRY
# wait = WebDriverWait(self.driver, 10)
# e = wait.until(EC.element_to_be_clickable((By.ID, "abx_register_country_chosen")))
# e.click()
# wait.until(EC.element_to_be_clickable((By.XPATH, "//li[text()='Malaysia']"))).click()
# OR
# self.driver.find_element(By.XPATH, "//div[@class='form-group abx-register-country']").click()
# self.driver.find_element(By.XPATH, "//li[contains(text(),'Malaysia')]").click()

# THE FIX = DROPDOWN 2 = MEMBERSHIP
# e = wait.until(EC.element_to_be_clickable((By.ID, "abx_register_subject_chosen")))
# e.click()
# wait.until(EC.element_to_be_clickable((By.XPATH, "//li[text()='Full Member']"))).click()

# THE FIX = DROPDOWN 3 = MEDNEFITS CREATE PROVIDER SITE. (Click Method)
# Right click on dropdown, inspect, copy rel x path, remove double quotes (Dont use selectors hub)
# Initial click code = //*[@id='service_list']/div[2]/div[1]/div[1]/div[1]/div/select"
# List of items code = //*[@id='service_list']/div[2]/div[1]/div[1]/div[1]/div/select/option[2]
# Change option number to select desired dropdown

# ----------------------------------------------------------------------------------------------
# Autosuggestive Dropdown = Dynamiclly select Particular Option from Autosuggestive Box using For-Loop

# driver.get("https://rahulshettyacademy.com/dropdownsPractise/")
# driver.find_element(By.XPATH, "//*[@id='autosuggest']").send_keys("ind")
# time.sleep(3)
#
# countries = driver.find_elements(By.CSS_SELECTOR, "li[class='ui-menu-item'] a")
# print(len(countries))
#
# for country in countries:
#     if country.text == "India":
#         print(country.text)
#         country.click()
#         break
#
# assert driver.find_element(By.XPATH, "//*[@id='autosuggest']").get_attribute("value") == "India"
# time.sleep(3)
# ---------------------------------------------------------------------------------------------------
# Selecting Checkboxes Dynamically (To be used if checkboxes position keeps changing/not known)

# driver.get("https://rahulshettyacademy.com/AutomationPractice/")
# Contruct your own Xpath if coding does not have, input ID, value ID, Name ID.
# Example : https://rahulshettyacademy.com/AutomationPractice/
# Selecting the entire list of "checkbox example" by using the type attribute (assuming the other ID's not available)
# Construct X Path Syntax = //input[@type='XXXXXXX']
# So the Syntax would be = //input[@type='checkbox']
# To verify whether the entire list is properly selected, input the constructed syntax in selectors hub and press enter.
# (cont) the number of elements found should match the number of checkboxes.

# checkboxes = driver.find_elements(By.XPATH, "//input[@type='checkbox']")
# print(len(checkboxes))
#
# for checkbox in checkboxes:
#     if checkbox.get_attribute("value") == "option2":
#         checkbox.click()
#         break
#
# time.sleep(3)
# -------------------------------------------------------------------------------------------------------
# Conditional Commands
# is.displayed = (Returns true/false whether element is displayed)
# is.enabled = (Returns true/false whether element is enabled)
# is.selected = (Returns true/false if element is selected)

# element1 = driver.find_element("name","userName" )
# print(element1.is_displayed())
# print(element1.is_enabled())
# print(element1.is_selected())

# element2 = driver.find_element("name", "password")
# print(element2.is_displayed())
# print(element2.is_enabled())
# print(element1.is_selected())

# ---------------------------------------------------------------------------------------------------------
# Assertion & Verify
# Asserts and Verify methods are commonly used in Selenium for verifying or validating tests/applications.
# Assert is case-sensitive

# Example 1
# driver.get("https://www.google.com.my/")
# assert "Google" in driver.title

# Example 2
# driver.get("https://rahulshettyacademy.com/angularpractice/")
# driver.find_element(By.XPATH, "/html/body/app-root/form-comp/div/form/input").click()
# time.sleep(2)
# message = driver.find_element(By.XPATH, "/html/body/app-root/form-comp/div/div[2]/div").text
# assert "Form" in message
# time.sleep(2)

# Example 3
# driver.get("https://rahulshettyacademy.com/angularpractice/")
# Email Column Box
# element1 = driver.find_element(By.XPATH, "//input[@name='email']")
# Check me out if you Love IceCreams Checkbox
# element2 = driver.find_element(By.XPATH, "//input[@id='exampleCheck1']")
# Element 1 Print & Assertion
# print(element1.is_enabled())
# print(element1.is_displayed())
# assert element1.is_enabled()
# assert element1.is_displayed()
# Element 2 Print & Assertion
# print(element2.is_selected())
# Assert not = Must return False value to pass test
# assert not element2.is_selected()

# MORE RESEARCH ON VARIOUS ASSERT METHODS REQUIRED
# https://www.softwaretestingmaterial.com/assertions-in-selenium-python/
# https://docs.pytest.org/en/7.1.x/how-to/assert.html#assertions-about-expected-exceptions
# https://understandingdata.com/posts/list-of-python-assert-statements-for-unit-tests/
# -----------------------------------------------------------------------------------------------------------
#  LESSON 1 = AUTOMATE THE FOLLOWING TEST CASE
# 1. Open the web Browser
# 2. Open URL ("https://admin-demo.nopcommerce.com/login")
# 3. Provide email & password (admin@yourstore.com : admin)
# 4. Click on login
# 5. Capture the title of the page once logged in
# 6. Verify the title of the page is same expected ("Dashboard / nopCommerce administration")
# 7. Close the browser

# Answer
# Open the Website
# driver.get("https://admin-demo.nopcommerce.com/login/")
# driver.implicitly_wait(10)

# Provide the email & Password
# driver.find_element(By.XPATH, "//*[@id='Email']").clear()
# driver.find_element(By.XPATH, "//*[@id='Email']").send_keys("admin@yourstore.com")
#
# driver.find_element(By.XPATH, "//*[@id='Password']").clear()
# driver.find_element(By.XPATH, "//*[@id='Password']").send_keys("admin")

# Click on Login
# driver.find_element(By.XPATH, "/html/body/div[6]/div/div/div/div/div[2]/div[1]/div/form/div[3]/button").click()
# time.sleep(5)

# Print the Title
# print(driver.title)

# Check if the Title Matches the expectation
# if (driver.title) == "Dashboard / nopCommerce administration":
#     print("Yes the Title is Correct")
# else:
#     print("No the title is different")
# time.sleep(5)
# driver.close()
# ---------------------------------------------------------------------------------------------------------
# How to handle Alerts & Pop-Ups

# driver.get("https://rahulshettyacademy.com/AutomationPractice/")
# Alert example/testing box

# driver.find_element(By.XPATH, "//*[@id='name']").send_keys("Krishna")
# driver.find_element(By.XPATH, "//*[@id='confirmbtn']").click()
# time.sleep(3)
# alert = driver.switch_to.alert

# Accept methods helps you to click on Pop up "Okay" button
# alert.accept()
# Dismiss methods helps you to click on Pop up "Cancel" button
# alert.dismiss()

# --------------------------------------------------------------
# How to Switch Windows & Tabs

# 1st Step = Capture Parent Window (not compulsory)
# parent_window = self.driver.current_window_handle

# 2nd Step = Object for all windows
# allwindows = self.driver.window_handles

# 3rd Step = Loop through windows & select the Correct Window
# for w in allwindows:
#     self.driver.switch_to.window(w)
#     if self.driver.title.__eq__("XXXXXXX XXXXX XXXXX"):
#         self.driver.find_element(By.XPATH, "XXXXX").click()
#         self.driver.find_element(By.XPATH, "XXXXX").click()
#         self.driver.find_element(By.XPATH, "XXXXX").click()
#         driver.close()
#         break

# 4th Step = Switch Back to Parent Window
# self.driver.switch_to_window(parent_window)

# ---------------------------------------------------------------------------------------------------------
# Mouse Interactions & Advanced Interactions (Open new window/tab,iFrames)
# ActionChains are a way to automate low level interactions such as mouse movements, mouse button actions, key press, and context menu interactions.
# When you call methods for actions on the ActionChains object, the actions are stored in a queue in the ActionChains object.
# When you call perform(), the events are fired in the order they are queued up.

# driver.get("https://rahulshettyacademy.com/AutomationPractice/")

# Required Base File
# action = ActionChains(driver)

# Scroll to element
# action.scroll_to_element(driver.find_element(By.XPATH, "//*[@id='mousehover']")).perform()
# time.sleep(3)

# Hover over element
# action.move_to_element(driver.find_element(By.XPATH, "//*[@id='mousehover']")).perform()
# time.sleep(3)

# Right click element
# action.context_click(driver.find_element(By.XPATH, "/html/body/div[4]/div/fieldset/div/div/a[1]")).perform()
# time.sleep(3)

# Left click element
# action.move_to_element(driver.find_element(By.XPATH, "/html/body/div[4]/div/fieldset/div/div/a[2]")).click().perform()
# time.sleep(3)

# ------------------------------------------------------------------------------------------------------------------
# iFrame in Selenium Webdriver is embedded HTML document that sits on top another HTML document. The iframe is often used to add content from other
# sources like an advertisement into a web page. When checking the HTML source code there should be an iFrame tag if website uses one.
# Use the switch command and input the frame name or frame id to switch

# driver.get("https://the-internet.herokuapp.com/iframe")
# driver.switch_to.frame("mce_0_ifr")
#
# driver.find_element(By.XPATH, "//*[@id='tinymce']/p").clear()
# driver.find_element(By.XPATH, "//*[@id='tinymce']/p").send_keys("Krishna is Learning!")
# time.sleep(5)

# Switching back to Normal window
# driver.switch_to.default_content()
# print(driver.find_element(By.XPATH, "//*[@id='content']/div/h3").text)
# -------------------------------------------------------------------------------------------------------------
# Javascript Executor/Using Javascript Commands to handle browser automation

# Window scrolling - some elements in website will only be available as you scroll down
# Javascript syntax = driver.execute_script()

# Scroll to the middle & bottom of the webpage, JS Events
# driver.get("https://rahulshettyacademy.com/AutomationPractice/")
# driver.execute_script("window.scrollBy(0,2000)")
# driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")
# time.sleep(7)

# Take screenshot of page & save in folder
# driver.get_screenshot_as_file("filename.png")

# Running headless Testing (Without browser opening), Ignore SSL errors, Default browser Check

# options = webdriver.ChromeOptions()
# options.add_argument("--headless")
# options.add_argument('--ignore-certificate-errors')
# options.add_argument("no-default-browser-check")
# options.add_argument("--start-maximized")
# options.add_argument("--disable-gpu")
# options.add_argument("--disable-notifications")

# service = Chrome()
# driver = webdriver.Chrome(service=service, options=options)

# driver.get("https://rahulshettyacademy.com/AutomationPractice/")
# print(driver.title)

# -----------------------------------------------------------------------------------------------------
# File Uploads & Downloads

# File Upload
# driver.get("https://testautomationpractice.blogspot.com/")
# driver.switch_to.frame("frame-one1434677811")
# file = "C://Users//Krishna//Desktop//Lanie1.JPG"
# driver.find_element(By.XPATH, "//*[@id='RESULT_FileUpload-10']").send_keys(file)
# time.sleep(8)

# File Download
# driver.get("https://demo.automationtesting.in/FileDownload.html")
# driver.find_element(By.XPATH, "//*[@id='pdfbox']").send_keys("I am learning File Downloads")
# driver.find_element(By.XPATH, "//*[@id='createPdf']").click()
# driver.find_element(By.XPATH, "//*[@id='pdf-link-to-download']").click()
# time.sleep(8)

# -----------------------------------------------------------------------------------------------------
# Data Driven Testing Using Excel

# Reading Data From Excel File -----------------------

# Create Path Folder
# path = "C:\\Users\\Krishna\\Desktop\\Software Testing\\sample_data.xlsx"

# Create Workbook Object & Store in Variable
# workbook = openpyxl.load_workbook(path)

# Extract the active sheet from the excel file
# sheet = workbook.active

# Get the number of rows & column on sheet
# rows = sheet.max_row
# cols = sheet.max_column
# print(rows)
# print(cols)


# Looping through the rows & columns using python Range Function(), the range function runs the loop until second last number only.
# Printing the Excel data to console & modify the loop make data presentable
# for r in range(1, rows+1):
#     for c in range(1, cols+1):
#         print(sheet.cell(row=r,column=c).value,end="   ")
#
#     print()

# Writing Data to Excel File ---------------------------

# Create Path Folder
# path = "C:\\Users\\Krishna\\Desktop\\Software Testing\\sample_data_empty.xlsx"

# Create Workbook Object & Store in Variable
# workbook = openpyxl.load_workbook(path)

# Extract the active sheet from the excel file
# sheet = workbook.active

# Write Data in Excel Sheets
# for r in range(1,5):
#     for c in range(1,5):
#         sheet.cell(row=r, column=c).value="krishna"

# Save the workbook
# workbook.save(path)

# Create Data driven Test Cases ------------------------------

# 1. Create common re-useable functions, normally stored in "utilities.py file"
# def getRowCount(file,sheetName):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     return(sheet.max_row)
#
# def getColumnCount(file,sheetName):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     return(sheet.max_column)
#
# def readData(file,sheetName,rownum, columnno):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     return sheet.cell(row=rownum, column=columnno).value
#
# def writeData(file,sheetName,rownum, columnno,data):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     sheet.cell(row=rownum, column=columnno).value = data
#     workbook.save(file)


# 2. Create Code to Run Tests
# driver.get("https://practicetestautomation.com/practice-test-login/")

# Store Excel Data in Variable
# path = "C:\\Users\\Krishna\\Desktop\\Software Testing\\data_driven_test2.xlsx"

# Call Re-Usable Functions
# rows = getRowCount(path, "Sheet1")

# Create loop to extract the data from file & input data into website (2 = Columns, Rows=Rows Function+1)
# for r in range(2,rows+1):
#     username = readData(path,"Sheet1",r,1)
#     password = readData(path,"Sheet1",r,2 )
#     time.sleep(3)
#
#     driver.find_element(By.XPATH, "//input[@id='username']").send_keys(username)
#     driver.find_element(By.XPATH, "//input[@id='password']").send_keys(password)
#     driver.find_element(By.XPATH, "//button[@id='submit']").click()
#     time.sleep(3)

# Verify login is successful by checking page title & send results to excel file & Log Out Button
#     if driver.title == "Logged In Successfully | Practice Test Automation":
#         print("Test Case Has Passed")
#         writeData(path,"Sheet1",r,3,"Test Passed")
#         driver.find_element(By.XPATH, "//a[normalize-space()='Log out']").click()
#     else:
#         print("Test Has Failed")
#         writeData(path, "Sheet1", r, 3, "Test Failed")

# Click home button again so loop will run again & input login parameters (Optional Based on Website)
#     driver.find_element(By.XPATH, "XXXXXXXXX").click()
#     time.sleep(3)

# ---------------------------------------------------------------------------------------------------------
# Browser Cookies in Testing

# 1. Capture all the Cookies created by Website
# driver.get("https://www.amazon.co.uk/")
# cookies = driver.get_cookies()
# print(cookies)
# print(len(cookies))

# 2. Adding new Cookie to browser & capture all the cookies
# cookie1 = {'name': 'Krishna', 'value': '12345'}
# driver.add_cookie(cookie1)

# cookies = driver.get_cookies()
# print(cookies)
# print(len(cookies))

# 3. Delete the Cookie
# cookie1 = {'name': 'Krishna', 'value': '12345'}
# driver.delete_cookie(cookie1)

# cookies = driver.get_cookies()
# print(cookies)
# print(len(cookies))

# -----------------------------------------------------------------------------------------------------
# Capture Screenshots in Selenium

# # Website & Scrolling Down
# driver.get("https://demo.guru99.com/test/newtours/")
# driver.execute_script("window.scrollBy(0,1000)")
#
# # 1. Screenshot Method & location to save Screenshot
# driver.save_screenshot("C:\\Users\\Krishna\\Desktop\\screenshot.png")

# ----------------------------------------------------------------------------------------------------
# Logging in Selenium (More on logging available in pytest Folder)

# 1. Create Basic log and Print to External file
# logging.basicConfig(filename="C:\\Users\\Krishna\\Desktop\\test.log",
#                     format=" %(asctime)s: %(levelname)s: %(message)s: ",
#                     level=logging.WARNING)

# logging.debug("This is a debug message")
# logging.info("This is a info Message")
# logging.warning("This is a warning message")
# logging.error("This is a error message")
# logging.critical("This is a critical message")

# ---------------------------------------------------------------------------------------------------
# Waits Commands in Selenium (Implicit Wait, Explicit Wait)
# Due to many websites using asynchronous javascript/ajax it is important to implement waiting mechanisms in many elements.

# Python code (time.sleep) will wait for the specified time regardless of whether its found something or not that's the difference
# time.sleep(10)

# General Wait, applies wait to all elements on the page (wait will remain provided there is no error in code), will move forward when element appears.
# Use driver.implicitly_wait(10)

# Explicit wait, wait applied for certain element & certain conditions to occur before proceeding further.
# Adding waits to specific lines of code & types of wait

# wait = WebDriverWait(driver, 20)

# wait.until(EC.visibility_of_element_located((By.XPATH, "XXXXXXX"))).click()
# wait.until(EC.presence_of_element_located((By.XPATH, "XXXXXXX"))).click()
# wait.until(EC.element_to_be_clickable((By.XPATH, "XXXXXXX"))).click()
# wait.until(EC.invisibility_of_element_located((By.XPATH, "XXXXX")))

# ----------------------------------------------------------------------------------------------------
# Create Dynamic X Path- Selenium/Python(Udemy) - https://omayo.blogspot.com/
# Click on the 5 links in the Site = compendiumdev, onlytestingblog, testwisely, jqueryui, theautomatedtester

# 1. Create 1 Dynamic X Path for selecting/clicking Multiple Elements
# - Select the main container & travel down - Verify X Path in Selectors HUB (No of elements) = //div[@id='LinkList1']//a
# - Wrap the entire X path in brackets & add index number = (//div[@id='LinkList1']//a)[1]
# - Change the index number to dynamic by adding the following = (//div[@id='LinkList1']//a)["+str(i)+"]
# - User For Loop =
# - for i in range(1,6):
#     dynamicxpath = "(//div[@id='LinkList1']//a)["+str(i)+"]"
#     driver.find_element(By.XPATH, dynamicxpath).click()
#     time.sleep(2)
#     driver.back()

# -----------------------------------------------------------------------------------------------------
# Handling Calendar Type 1 =  Test Calendar with a Function & just call the function with the Specific Dates

# # Calendar Website
# driver.get("https://practice-automation.com/calendars/")
#
# # Click to open the calendar
# driver.find_element(By.XPATH, "//input[@data-format='yy-mm-dd']").click()
#
# # Wait for the calendar to appear
# wait = WebDriverWait(driver, 10)
# wait.until(EC.visibility_of_element_located((By.XPATH, "//*[@id='ui-datepicker-div']")))
#
#
# def calendar_test(date, month, year, direction):
#     # Save month & Year into a Variable
#     current_month = driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/div/span[1]").text
#     current_year = driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/div/span[2]").text
#
#     # Use the "while not" loop to select the month & year
#     while not (current_month == month and current_year == year):
#         if direction == "back":
#             driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/a[1]").click()
#         elif direction == "forward":
#             driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/a[2]").click()
#         current_month = driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/div/span[1]").text
#         current_year = driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/div/span[2]").text
#
#     # Select the Date
#     dynamic_xpath = "//td[@data-handler='selectDay']/a[text()='"+date+"']"
#     driver.find_element(By.XPATH, dynamic_xpath).click()
#
#
# calendar_test("15", "February", "2028", "forward")

# Handling Calendar Type 1 =  Using Javascript (EASY) -------------------------------------------

# # Calendar Website
# driver.get("https://seleniumpractise.blogspot.com/2016/08/how-to-handle-calendar-in-selenium.html")
#
# driver.execute_script("document.getElementById('datepicker').value='29/08/2028'")
# time.sleep(6)

# --------------------------------------------------------------------------------------------------------------------
# Version Control Software (GIT/GITHUB)

# Popular Version Control Softwares
# 1. Git(Software Name), GitHub(Place where repository is stored)
# 2. SVN (Apache Subversion)
# 3. Concurrent Version System (CVS)

# All Basic Git Commands
# https://confluence.atlassian.com/bitbucketserver/basic-git-commands-776639767.html

# Git & GitHub Workflow





















# ---------------------------------------------------------------------------------------------------
# Jenkins - Free Open Source Continous Intergration Server
# For full course Refer to RVCA/UDEMY

# Jenkins is an open source continuous integration/continuous delivery and deployment (CI/CD) automation software, It helps automate
# the parts of software development related to building, testing, and deploying, facilitating continuous integration and continuous delivery.
# It is a server-based system that runs in servlet containers such as Apache Tomcat.

# What is CI/CD?
# CI and CD stands for continuous integration and continuous delivery/continuous deployment. In very simple terms,
# (con't) CI is a modern software development practice in which incremental code changes are made frequently and reliably.
# (con't)Automated build-and-test steps triggered by CI ensure that code changes being merged into the repository are reliable.
# (con't)The code is then delivered quickly and seamlessly as a part of the CD process. In the software world, the CI/CD pipeline
# (con't) refers to the automation that enables incremental code changes from developers’ desktops to be delivered quickly and reliably to production.

# What is the difference between CI and CD?
# Continuous integration (CI) is practice that involves developers making small changes and checks to their code.
# Due to the scale of requirements and the number of steps involved, this process is automated to ensure that teams can build, test, and package
# their applications in a reliable and repeatable way. CI helps streamline code changes, thereby increasing time for developers to make changes
# and contribute to improved software.

# Continuous delivery (CD) is the automated delivery of completed code to environments like testing and development. CD provides an automated and consistent
# way for code to be delivered to these environments. Continuous deployment is the next step of continuous delivery. Every change that passes the automated
# tests is automatically placed in production, resulting in many production deployments. Continuous deployment should be the goal of most companies that
# are not constrained by regulatory or other requirements. In short, CI is a set of practices performed as developers are writing code,
# and CD is a set of practices performed after the code is completed.

# -----------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------- PYTHON FUNDAMENTALS -----------------------------------------------------------------

# PYTHON STRING METHODS

# STRING SLICING
# message1 = "Success! Transaction MY-04786-0005806 has been successfully created"
# message2 = message1[7]
# print(message2) = !

# message1 = "Success! Transaction MY-04786-0005806 has been successfully created"
# message2 = message1[10]
# print(message2) = r

# message1 = "Success! Transaction MY-04786-0005806 has been successfully created"
# message2 = message1[-7]
# print(message2) = c

# message1 = "Success! Transaction MY-04786-0005806 has been successfully created"
# message2 = message1[-10]
# print(message2) = l

# message1 = "Success! Transaction MY-04786-0005806 has been successfully created"
# message2 = message1[10:-10]
# print(message2) = ransaction MY-04786-0005806 has been successful

# message1 = "Success! Transaction MY-04786-0005806 has been successfully created"
# message2 = message1[20:-20]
# print(message2) = MY-04786-0005806 has been

# message1 = "Success! Transaction MY-04786-0005806 has been successfully created"
# message2 = message1[20:-30]
# print(message2) = MY-04786-0005806

# -----------------------------------------------
# Convert String With Multiple Lines to Single Line

# Original String = viamednefits1
# "MYR 120.00
# (Cap per visit: No Limit)"

# 1st Step - Convert to List
# viamednefits2 = viamednefits1.split()
# Output = ['MYR', '120.00', '(Cap', 'per', 'visit:', 'No', 'Limit)']

# 2nd Step - Capture the index we want in the list
# viamednefits3 = viamednefits2[0:2]
# Output = ['MYR', '120.00']

# 3rd Step = Convert List to normal String
# converted_list = map(str, viamednefits3)
# viamednefits4 = ''.join(converted_list)
# Output = MYR120.00

# -------------------------------------------------
# Modify Characters/Adding Space/Add Signs between Characters in String

# viamednefits4 = "MYR120.00"
# viamednefits5 = viamednefits4[:3] + "++" + viamednefits4[3:]
# print(viamednefits5)
# Output = MYR++120.00

# viamednefits4 = "MYR120.00"
# viamednefits5 = viamednefits4[:3] + " " + viamednefits4[3:]
# print(viamednefits5)
# Output = MYR 120.00

# viamednefits4 = "MYR120.00"
# viamednefits5 = "++" + " " + viamednefits4
# print(viamednefits5)
# Output = ++ MYR 120.00

# viamednefits4 = "MYR120.00"
# viamednefits5 = viamednefits4 + " " + "++"
# print(viamednefits5)
# Output = MYR120.00 ++

# ---------------------------------------------------
# Python TypeCasting - Convert Float to Integer, String to Integer, Integer to String

# Float to Integer
# a = 10.8
# print(int(a))
# Output = 10

# Integer to Float
# a = 10
# print(float(a))
# Output = 10.0

# Integer to String
# a = 10
# print(type(a))
# b = str(a)
# print(type(b))
# Output = "10" as String

# String To Integer
# a = "10"
# print(type(a))
# b = int(a)
# print(type(b))
# Output = "10" as Integer
# -----------------------------------------------------
# Python replace string with Variable input & Indexing the Input

# name = "krishna"
# experience = 20
# location = "Subang Jaya"
# print("Hello my name is " + name + " and i have " +str(experience) +  " years of experience")

# Without Indexing
# name = "krishna"
# experience = 20
# location = "Subang Jaya"
# print("Hello my name is {}, i have {} years of experience & i live in {}".format(name, experience, location))


# With Indexing
# name = "krishna"
# experience = 20
# location = "Subang Jaya"

# Index
# experience is index 0
# location is index 1
# name is index 2
# print("Hello my name is {2}, i have {0} years of experience & i live in {1}".format(experience, location, name))
# ---------------------------------------------------------------------------------------------------------------
# PYTHON LOOPS

# While Lopp (Clicking on same button multiple times (12 times) in a row by using "Countered While Loop")

# variablename = 1

# while variablename <= 12:
#     self.driver.find_element(By.XPATH, "XXXXXXXXXXX").click()
#     variablename = variablename + 1


# For Loop With Range() - Clicking on button 10 Times

# for i in range(10):
#     driver.find_element(By.XPATH, "XXXXXXXXX").click()

# for i in range(1,11):
#     driver.find_element(By.XPATH, "XXXXXXXXX").click()

# ---------------------------------------------------------------------------------------------------------------
# PYTHON (TRANSFER/JUMP STATEMENTS = Break, Continue)

# Break Statement = "Stops the entire loop"

# for i in range(10):
#     if i == 8:
#         break
#     print(i)


# Continue Statement = "Cancel's only the current iteration"

# for i in range(10):
#     if i == 8:
#         continue
#     print(i)

# --------------------------------------------------------------------------------------------------------------
# PYTHON ERROR HANDLING (TRY, CATCH & EXCEPT STATEMENTS)

# A Try-Except statement is a code block that allows your program to take alternative actions in case an error occurs, even if there
# an error in the code, it will catch it, perform any alternative action and move on to execute the remaining part of the code. Without the
# try/catch statement the program will fail immediately is there is an error.

# Examples:

# a = 10
# b = 0

# try:
#     print(a/b)
# except Exception:
#     print("Hey, you cannot divide a number by zero")

# print("This line of code is still working")

# ----------------------------------------------------------------
# Finally - The finally block, if specified, will be executed regardless if the try block raises an error or not

# try:
#   print(x)
# except:
#   print("Something went wrong")
# finally:
#   print("The 'try except' is finished")

# -----------------------------------------------------------------------------------------------------------
# Practice QAFOX - Udemy Course

# driver.get("https://omayo.blogspot.com/")







